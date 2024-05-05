import pandas as pd
import numpy as np
from tkinter import filedialog
import time
import os
from tkinter.messagebox import showinfo
try:
    import LIB.formatos as fmt
except:
    import formatos as fmt
import openpyxl

def Consolidador_Excel():

    StartTime = time.time()

    # #crear una funcion para seleccionar todos los archivos .xlsx de una carpeta
    # def seleccionar_carpeta():

    # Abre una ventana para seleccionar la carpeta
    ruta = filedialog.askopenfile(title="Seleccionar el Excel que posee los Directorios con los archivos a consolidar" , filetypes = (("Excel files","*.xlsx"),("all files","*.*")) )

    archivos = pd.read_excel(ruta.name)

    # Hacer una lista con el la primer columna del excel
    archivos = archivos.iloc[:,0].tolist()

    # Eliminar los valores nulos de la lista y los que no terminan en .xlsx
    archivos = [f for f in archivos if str(f) != 'nan']

    # Reemplazar los \ por / en la lista
    archivos = [f.replace("\\", "/") for f in archivos]

    # Consolidar Archivos
    Consolidador(archivos)

    EndTime = time.time()
    Ejecucion = EndTime - StartTime
    Ejecucion = round(Ejecucion, 2)

    # Crear una ventana de mensaje con "El archivo se ha consolidado correctamente en Segundo"
    showinfo("Consolidador" , "El archivo se ha consolidado correctamente en " + str(Ejecucion) + " segundos")


def Consolidador_Carpetas():
    '''
    Consolidar archivos de Excel de Mis Comprobantes en Base a una carpeta seleccionada
    '''

    Carpeta = filedialog.askdirectory(title="Seleccionar la carpeta que posee los archivos a consolidar")

    StartTime = time.time()

    # Listar archivos de la carpeta seleccionada
    archivos = os.listdir(Carpeta)

    # Filtrar archivos de Excel
    archivos = [f for f in archivos if f.endswith('.xlsx')]

    # Agregar la ruta de la carpeta a cada archivo
    archivos = [Carpeta + "/" + f for f in archivos]

    # Consolidar Archivos
    Consolidador(archivos)

    EndTime = time.time()
    Ejecucion = EndTime - StartTime
    Ejecucion = round(Ejecucion, 2)

    # crear un mensaje con el tiempo de ejecución
    showinfo("Consolidador" , "El archivo se ha consolidado correctamente en " + str(Ejecucion) + " segundos")



def Consolidador(archivos: list):
    '''
    Consolidar archivos de Excel de Mis Comprobantes en Base a una lista de archivos
    
    Parameters
    ----------
    archivos : list
        Lista de archivos de Excel de Mis Comprobantes
    '''

    TablaBase = pd.DataFrame()

    # Consolidar archivos y renombrar columnas
    # consolidadar columnas
    for f in archivos:

        try:

            #Si el existe el archivo, leerlo
            if os.path.isfile(f):  
                data = pd.read_excel(f, header = None, skiprows=2 , )
                # si el datsaframe esta vacio, no hacer nada
                if len(data) > 0:

                    # Crear la columna 'Archivo' con el ultimo elemento de 'f' separado por "/"
                    data['Archivo'] = f.split("/")[-1]
                    #data['Archivo'] = f.str.split("/")[-1]
                    data['CUIT Cliente'] = data["Archivo"].str.split("-").str[3].str.strip().astype(np.int64)
                    data['Fin CUIT'] = data["Archivo"].str.split("-").str[0].str.strip().astype(np.int64)
                    TablaBase = pd.concat([TablaBase , data])

        except:
            pass
            
    # Renombrar columnas
    TablaBase.columns = [ 'Fecha' , 'Tipo' , 'Punto de Venta' , 'Número Desde' , 'Número Hasta' , 'Cód. Autorización' , 'Tipo Doc. Receptor' , 'Nro. Doc. Receptor/Emisor' , 'Denominación Receptor/Emisor' , 'Tipo Cambio' , 'Moneda' , 'Imp. Neto Gravado' , 'Imp. Neto No Gravado' , 'Imp. Op. Exentas' , 'Otros Tributos' , 'IVA' , 'Imp. Total' , 'Archivo' , 'CUIT Cliente' , 'Fin CUIT']

    #Multiplicar por tipo de cambio
    TablaBase['Imp. Neto Gravado'] *= TablaBase['Tipo Cambio']
    TablaBase['Imp. Neto No Gravado'] *= TablaBase['Tipo Cambio']
    TablaBase['Imp. Op. Exentas'] *= TablaBase['Tipo Cambio']
    TablaBase['IVA'] *= TablaBase['Tipo Cambio']
    TablaBase['Imp. Total'] *= TablaBase['Tipo Cambio']
    TablaBase['Otros Tributos'] *= TablaBase['Tipo Cambio']   

    TablaBase['Tipo'] = TablaBase['Tipo'].str.strip().str.upper()
    
    #Cambiar de signo si es una Nota de Crédito
    TablaBase.loc[TablaBase["Tipo"].str.contains("NOTA DE CRÉDITO"), ['Imp. Neto Gravado' , 'Imp. Neto No Gravado' , 'Imp. Op. Exentas' , 'Otros Tributos', 'IVA' , 'Imp. Total']] *= -1

    #Crear columna de 'MC' con los valores 'archivo' que van desde el caracter 5 al 8 en la TablaBase
    TablaBase['MC'] = TablaBase['Archivo'].str.split("-").str[1].str.strip()

    #Crear Tabla dinámica con los totales de las columnas  'Imp. Neto Gravado' , 'Imp. Neto No Gravado' , 'Imp. Op. Exentas' , 'IVA' , 'Imp. Total' por 'Archivo'
    TablaDinamica = pd.pivot_table(TablaBase, values=['Imp. Neto Gravado' , 'Imp. Neto No Gravado' , 'Imp. Op. Exentas' , 'IVA' , 'Imp. Total' , 'Tipo'], index=['Archivo'], aggfunc={'Imp. Neto Gravado': np.sum , 'Imp. Neto No Gravado': np.sum , 'Imp. Op. Exentas': np.sum , 'IVA': np.sum , 'Imp. Total': np.sum , 'Tipo': 'count'})

    #Crear Tabla dinámica con los totales de las columnas  'Imp. Neto Gravado' , 'Imp. Neto No Gravado' , 'Imp. Op. Exentas' , 'IVA' , 'Imp. Total' por 'CUIT Cliente'
    TablaDinamica2 = pd.pivot_table(TablaBase, values=['Imp. Neto Gravado' , 'Imp. Neto No Gravado' , 'Imp. Op. Exentas' , 'IVA' , 'Imp. Total'], index=['CUIT Cliente' , 'MC' , 'Archivo' , 'Tipo'], aggfunc={'Imp. Neto Gravado': np.sum , 'Imp. Neto No Gravado': np.sum , 'Imp. Op. Exentas': np.sum , 'IVA': np.sum , 'Imp. Total': np.sum , 'Tipo': 'count'})

    #Crear Tabla dinámica con los totales de las columnas  'Imp. Neto Gravado' , 'Imp. Neto No Gravado' , 'Imp. Op. Exentas' , 'IVA' , 'Imp. Total' por 'CUIT Cliente' , 'MC' , 'Archivo' y 'Tipo'
    #TablaDinamica3 = pd.pivot_table(TablaBase, values=['Imp. Neto Gravado' , 'Imp. Neto No Gravado' , 'Imp. Op. Exentas' , 'IVA' , 'Imp. Total'], index=['CUIT Cliente' , 'MC' , 'Archivo' , 'Tipo'], aggfunc=np.sum)
    #Crear ua tabla dinámica como la anterior pero agregándo la cantidad de registros que conforman el tipo
    #TablaDinamica3 = pd.pivot_table(TablaBase, values=['Imp. Neto Gravado' , 'Imp. Neto No Gravado' , 'Imp. Op. Exentas' , 'IVA' , 'Imp. Total'], index=['CUIT Cliente' , 'MC' , 'Archivo' , 'Tipo'], aggfunc={'Imp. Neto Gravado' : np.sum , 'Imp. Neto No Gravado' : np.sum , 'Imp. Op. Exentas' : np.sum , 'IVA' : np.sum , 'Imp. Total' : np.sum , 'Tipo' : 'count'})

    # Renombrar la columna 'Tipo' por 'Cantidad de Comprobantes' de la TablaDinamica1 , TablaDinamica2 y TablaDinamica3
    TablaDinamica.rename(columns={'Tipo': 'Cantidad de Comprobantes'}, inplace=True)
    TablaDinamica2.rename(columns={'Tipo': 'Cantidad de Comprobantes'}, inplace=True)
    #TablaDinamica3.rename(columns={'Tipo': 'Cantidad de Comprobantes'}, inplace=True)

    # Exportar
    with pd.ExcelWriter('Consolidado.xlsx') as Archivo_final:
        #Exportar Tabla Base a la hoja 'Consolidado' de 'Consolidado.xlsx'
        TablaBase.to_excel(Archivo_final, sheet_name="Consolidado" , index=False)
        #Exportar Tabla Dinámica a la hoja 'TD' de 'Consolidado.xlsx'
        TablaDinamica.to_excel(Archivo_final, sheet_name="TD" , index=True , merge_cells=False)
        #Exportar Tabla Dinámica a la hoja 'TD2' de 'Consolidado.xlsx'
        TablaDinamica2.to_excel(Archivo_final, sheet_name="TD Cruce" , index=True , merge_cells=False)

    # Aplicar formatos
    workbook = openpyxl.load_workbook('Consolidado.xlsx')
    hoja1 = workbook['Consolidado']  # Nombre de la hoja del DataFrame
    hoja2 = workbook['TD']  # Nombre de la hoja del DataFrame
    hoja3 = workbook['TD Cruce']  # Nombre de la hoja del DataFrame

    Hojas = [hoja1 , hoja2 , hoja3]

    for hoja in Hojas:    
        fmt.Aplicar_formato_encabezado(hoja)
        fmt.Autoajustar_columnas(hoja)
        fmt.Agregar_filtros(hoja)

    fmt.Aplicar_formato_moneda(hoja1 , 10 , 16)
    fmt.Aplicar_formato_moneda(hoja2 , 2 , 6)
    fmt.Aplicar_formato_moneda(hoja3 , 5 , 9)

    fmt.Alinear_columnas(hoja2 , 1 , 1 , 'left')
    fmt.Alinear_columnas(hoja3 , 1 , 4 , 'left')

    # Guardar el archivo Excel
    workbook.save('Consolidado.xlsx')

if __name__ == "__main__":
    Consolidador_Carpetas()