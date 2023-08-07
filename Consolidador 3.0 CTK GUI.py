import pandas as pd
import numpy as np
import os
import customtkinter as ctk
from tkinter import filedialog
import openpyxl
from openpyxl.styles import PatternFill, Font , Alignment
from openpyxl.worksheet.worksheet import Worksheet



###### TKinter #############################################

# Crea la ventana principal
ventana = ctk.CTk()
ventana.title("Consolidador 3.0")
ventana.geometry("400x200")

# Crea una etiqueta para mostrar la lista de archivos
etiqueta = ctk.CTkLabel(ventana, text="")
etiqueta.pack()

#crear una funcion para seleccionar todos los archivos .xlsx de una carpeta
def seleccionar_carpeta():

    # Abre una ventana para seleccionar la carpeta
    ruta = filedialog.askdirectory(parent=ventana)

    # Si se seleccionó una carpeta
    if ruta != "":

        # Obtiene la lista de archivos .xlsx de la carpeta
        archivos = os.listdir(ruta)
        archivos = [x for x in archivos if x.endswith(".xlsx")]
        
        #Crear un textbox con los archivos consolidados
        Archivos_Textbox = ctk.CTkTextbox(ventana, width=40, height=10)
        Archivos_Textbox.insert(ctk.END, f"Archivos consolidados: {str(len(archivos))}\n")
        for a in archivos:
          Archivos_Textbox.insert(ctk.END, f"{a}\n")
        Archivos_Textbox.pack(fill=ctk.BOTH, expand=True , pady=10)
        
        # Si hay archivos
        if len(archivos) > 0:

            # Muestra la lista de archivos en la etiqueta
            #etiqueta["text"] = "Procesando " + str(len(archivos)) + " archivos .xlsx"

            # Crear tabla a donde consolidar
            TablaBase = pd.DataFrame()

            # Consolidar archivos y renombrar columnas
            # consolidadar columnas
            for f in archivos:
                data = pd.read_excel(ruta + "/" + f, header = None, skiprows=2 , )
                # si el datsaframe esta vacio, no hacer nada
                if len(data) > 0:
                    data['Archivo'] = f
                    data['CUIT Cliente'] = data["Archivo"].str.split("-").str[3].str.strip().astype(np.int64)
                    data['Denominación Cliente'] = data["Archivo"].str.split("-").str[4].str.strip()
                    data['Fin CUIT'] = data["Archivo"].str.split("-").str[0].str.strip().astype(np.int64)
                    TablaBase = pd.concat([TablaBase , data])
                
            # Renombrar columnas
            TablaBase.columns = [ 'Fecha' , 'Tipo' , 'Punto de Venta' , 'Número Desde' , 'Número Hasta' , 'Cód. Autorización' , 'Tipo Doc. Receptor' , 'Nro. Doc. Receptor/Emisor' , 'Denominación Receptor/Emisor' , 'Tipo Cambio' , 'Moneda' , 'Imp. Neto Gravado' , 'Imp. Neto No Gravado' , 'Imp. Op. Exentas' , 'IVA' , 'Imp. Total' , 'Archivo' , 'CUIT Cliente' , 'Denominación Cliente' , 'Fin CUIT']

            #Multiplicar por tipo de cambio
            TablaBase['Imp. Neto Gravado'] *= TablaBase['Tipo Cambio']
            TablaBase['Imp. Neto No Gravado'] *= TablaBase['Tipo Cambio']
            TablaBase['Imp. Op. Exentas'] *= TablaBase['Tipo Cambio']
            TablaBase['IVA'] *= TablaBase['Tipo Cambio']
            TablaBase['Imp. Total'] *= TablaBase['Tipo Cambio']   

            #Cambiar de signo si es una Nota de Crédito
            TablaBase.loc[TablaBase["Tipo"].str.contains("Nota de Crédito"), ['Imp. Neto Gravado' , 'Imp. Neto No Gravado' , 'Imp. Op. Exentas' , 'IVA' , 'Imp. Total']] *= -1
            
            #Crear columna de 'MC' con los valores 'archivo' que van desde el caracter 5 al 8 en la TablaBase
            TablaBase['MC'] = TablaBase['Archivo'].str.split("-").str[1].str.strip()

            #Crear Tabla dinámica con los totales de las columnas  'Imp. Neto Gravado' , 'Imp. Neto No Gravado' , 'Imp. Op. Exentas' , 'IVA' , 'Imp. Total' por 'Archivo'
            TablaDinamica = pd.pivot_table(TablaBase, values=['Imp. Neto Gravado' , 'Imp. Neto No Gravado' , 'Imp. Op. Exentas' , 'IVA' , 'Imp. Total' , 'Tipo'], index=['Archivo'], aggfunc={'Imp. Neto Gravado': np.sum , 'Imp. Neto No Gravado': np.sum , 'Imp. Op. Exentas': np.sum , 'IVA': np.sum , 'Imp. Total': np.sum , 'Tipo': 'count'})

            #Crear Tabla dinámica con los totales de las columnas  'Imp. Neto Gravado' , 'Imp. Neto No Gravado' , 'Imp. Op. Exentas' , 'IVA' , 'Imp. Total' por 'CUIT Cliente'
            TablaDinamica2 = pd.pivot_table(TablaBase, values=['Imp. Neto Gravado' , 'Imp. Neto No Gravado' , 'Imp. Op. Exentas' , 'IVA' , 'Imp. Total'], index=['CUIT Cliente' , 'MC' , 'Archivo' , 'Tipo'], aggfunc={'Imp. Neto Gravado': np.sum , 'Imp. Neto No Gravado': np.sum , 'Imp. Op. Exentas': np.sum , 'IVA': np.sum , 'Imp. Total': np.sum , 'Tipo': 'count'})

            #Crear Tabla dinámica con los totales de las columnas  'Imp. Neto Gravado' , 'Imp. Neto No Gravado' , 'Imp. Op. Exentas' , 'IVA' , 'Imp. Total' por 'CUIT Cliente' , 'MC' , 'Archivo' y 'Tipo'
            #TablaDinamica3 = pd.pivot_table(TablaBase, values=['Imp. Neto Gravado' , 'Imp. Neto No Gravado' , 'Imp. Op. Exentas' , 'IVA' , 'Imp. Total'], index=['CUIT Cliente' , 'MC' , 'Archivo' , 'Tipo'], aggfunc=np.sum)
            #Crear ua tabla dinámica como la anterior pero agregándo la cantidad de registros que conforman el tipo
            #TablaDinamica3 = pd.pivot_table(TablaBase, values=['Imp. Neto Gravado' , 'Imp. Neto No Gravado' , 'Imp. Op. Exentas' , 'IVA' , 'Imp. Total'], index=['CUIT Cliente' , 'MC' , 'Archivo' , 'Tipo'], aggfunc={'Imp. Neto Gravado' : np.sum , 'Imp. Neto No Gravado' : np.sum , 'Imp. Op. Exentas' : np.sum , 'IVA' : np.sum , 'Imp. Total' : np.sum , 'Tipo' : 'count'})

            # Renombrar la columna 'Tipo' por 'Cantidad de Comprobantes' de la TablaDinamica1 , TablaDinamica2 y TablaDinamica3
            TablaDinamica.rename(columns={'Tipo': 'Cantidad de Comprobantes'}, inplace=True)
            TablaDinamica2.rename(columns={'Tipo': 'Cantidad de Comprobantes'}, inplace=True)
            #TablaDinamica3.rename(columns={'Tipo': 'Cantidad de Comprobantes'}, inplace=True)

            # Exportar
            with pd.ExcelWriter('Consolidado.xlsx') as Archivo_final:

                TablaBase.to_excel(Archivo_final, sheet_name="Consolidado" , index=False)

                #Exportar Tabla Dinámica a la hoja 'TD' de 'Consolidado.xlsx'
                TablaDinamica.to_excel(Archivo_final, sheet_name="TD" , index=True , merge_cells=False)

                #Exportar Tabla Dinámica a la hoja 'TD2' de 'Consolidado.xlsx'
                TablaDinamica2.to_excel(Archivo_final, sheet_name="TD Cruce" , index=True , merge_cells=False)

            #Exportar Tabla Dinámica a la hoja 'TD por CBTE' de 'Consolidado.xlsx'
            #TablaDinamica3.to_excel(Archivo_final, sheet_name="TD por CBTE" , index=True , merge_cells=False)
            
            workbook = openpyxl.load_workbook('Consolidado.xlsx')
            hoja1 = workbook['Consolidado']  # Nombre de la hoja del DataFrame
            hoja2 = workbook['TD']  # Nombre de la hoja del DataFrame
            hoja3 = workbook['TD Cruce']  # Nombre de la hoja del DataFrame

            # Darle formato a los Títulos de las columnas
            Fondotitulo = PatternFill(start_color='002060' , end_color='002060' ,  fill_type='solid')
            LetraColor = Font(color='FFFFFF')

            # Funciones para formatear Excel

            # Aplicar formato al encabezado
            def Aplicar_formato_encabezado(HojaActual : Worksheet):
                '''
                Función que aplica formato al encabezado de la hoja
                '''
                
                for cell in HojaActual[1]:
                    cell.fill = Fondotitulo
                    cell.font = LetraColor


            # Aplica formato de moneda a las columnas de importes
            def Aplicar_formato_moneda(HojaActual : Worksheet , ColumnaInicial : int , ColumnaFinal : int):
                '''
                Función que aplica formato de moneda a las columnas de importes
                '''
                
                formato = '_-* #,##0.00_-;-* #,##0.00_-;_-* "-"??_-;_-@_-'

                for cell in HojaActual.iter_rows(min_row=2, min_col=ColumnaInicial, max_row=HojaActual.max_row, max_col=ColumnaFinal):
                    for celda in cell:
                        celda.number_format = formato


            # Autoajustar los anchos de las columnas según el contenido
            def Autoajustar_columnas(HojaActual : Worksheet):
                '''
                Función que autoajusta las columnas de la hoja
                '''
                
                for column_cells in HojaActual.columns:
                    length = max(len(str(cell.value)) for cell in column_cells)
                    HojaActual.column_dimensions[column_cells[0].column_letter].width = length + 2


            # Agregar filtros de datos a las hojas
            def Agregar_filtros(HojaActual : Worksheet):
                '''
                Función que agrega filtros a la hoja
                '''
                
                HojaActual.auto_filter.ref = HojaActual.dimensions

            # Alinear columnas
            def Alinear_columnas(HojaActual : Worksheet , ColumnaInicial : int , ColumnaFinal : int , Alineacion : str):
                '''
                Función que alinea las columnas de la hoja
                '''
                
                for cell in HojaActual.iter_rows(min_row=2, min_col=ColumnaInicial, max_row=HojaActual.max_row, max_col=ColumnaFinal):
                    for celda in cell:
                        celda.alignment = Alineacion



            # Aplicar formatos
            Aplicar_formato_encabezado(hoja1)
            Aplicar_formato_encabezado(hoja2)
            Aplicar_formato_encabezado(hoja3)

            Aplicar_formato_moneda(hoja1 , 10 , 16)
            Aplicar_formato_moneda(hoja2 , 2 , 6)
            Aplicar_formato_moneda(hoja3 , 5 , 9)
            
            Autoajustar_columnas(hoja1)
            Autoajustar_columnas(hoja2)
            Autoajustar_columnas(hoja3)

            Agregar_filtros(hoja1)
            Agregar_filtros(hoja2)
            Agregar_filtros(hoja3)

            # Alinar toda la columna A de la hoja 2 a la izquierda
            Alinear_columnas(hoja2 , 1 , 1 , Alignment(horizontal='left'))
            Alinear_columnas(hoja3 , 1 , 4 , Alignment(horizontal='left'))

            # Guardar el archivo Excel
            workbook.save('Consolidado.xlsx')

        else:
            
            # Muestra un mensaje de error en la etiqueta
            etiqueta["text"] = "No hay archivos .xlsx en la carpeta seleccionada"

        

#Crea un botón para seleccionar la carpeta desde una ventana del explorador de archivos

boton = ctk.CTkButton(ventana, text="Seleccionar carpeta", command=seleccionar_carpeta)
boton.pack()

#Crear un botón para salir de la aplicación al lado del botón anterior
boton2 = ctk.CTkButton(ventana, text="Salir", command=ventana.quit)
boton2.pack(pady=10)



# Inicia el bucle principal de la ventana
ventana.mainloop()